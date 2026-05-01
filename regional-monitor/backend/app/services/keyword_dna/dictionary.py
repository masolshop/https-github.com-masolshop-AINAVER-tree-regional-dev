"""사전 빌더 — 6 카테고리 분류 + 회선수 가중치 저장.

카테고리:
  - main:     중심 키워드 (하수구, 흥신소, 누수, 보일러, 열쇠, 에어컨 …)
  - place:    장소/대상 (사무실, 가정, 아파트, 상가, 공장, 화장실 …)
  - action:   동작/서비스 (막힘, 뚫음, 설치, 수리, 교체, 청소, 출장, 매입, 판매 …)
  - material: 재료/원인 (변기, 싱크대, 수도, 배관, 도어락, 폐기물 …)
  - brand:    브랜드 (LG, 삼성, KCC, 린나이, 경동, 귀뚜라미, 엘지, 하우시스 …)
  - tag:      수식어/태그 (24시, 무료, 전문, 업체, 센터, 사무, 당일, 긴급, 1급 …)

오프라인 1회 빌드 후 dictionary.json 으로 캐시.
"""
from __future__ import annotations

import json
import os
import re
from collections import Counter, defaultdict
from functools import lru_cache
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

CATEGORIES = ["main", "action", "material", "place", "brand", "tag"]

DATA_DIR = Path(__file__).resolve().parents[2] / "data" / "keyword_dna"
BUSINESS_XLSX = DATA_DIR / "business_names.xlsx"
CATEGORY_XLSX = DATA_DIR / "categories.xlsx"
DICTIONARY_JSON = DATA_DIR / "dictionary.json"


# ──────────────────────────────────────────────────────────────────────────
# 시드 사전 — 카테고리별 핵심 어휘 (수동 큐레이션, 약 250개)
# ──────────────────────────────────────────────────────────────────────────
SEED: Dict[str, List[str]] = {
    "main": [
        "하수구", "변기", "싱크대", "수도", "배관", "보일러", "에어컨",
        "흥신소", "심부름센터", "탐정", "사설조사",
        "열쇠", "도어락", "도어", "번호키",
        "폐기물", "쓰레기", "유품", "고철", "고물", "이삿짐", "이사",
        "유리", "거울", "샷시", "샤시", "창호", "창문", "방충망",
        "CCTV", "컴퓨터", "노트북", "PC",
        "꽃", "화환", "근조",
        "스카이", "사다리차", "크레인", "포크레인", "지게차",
        "중고차", "자동차", "타이어",
        "현수막", "간판", "블라인드", "커튼",
        "장례식", "장례", "결혼식",
        "유선방송", "위성방송",
        "가스", "전기", "통신",
        "용달", "퀵", "택배", "운전대행", "대리운전",
    ],
    "action": [
        "막힘", "뚫음", "뚫", "역류", "해빙", "동파", "누수", "탐지",
        "설치", "수리", "교체", "공사", "시공", "철거", "복구", "복귀",
        "청소", "정리", "수거", "처리", "매입", "판매", "임대", "렌탈",
        "출장", "배달", "이전", "이송",
        "조사", "추적", "찾기", "감시", "미행",
        "상담", "컨설팅", "견적", "AS",
        "고압세척", "방수", "포맷", "조립",
        "관리", "검사", "점검", "측정",
    ],
    "material": [
        "변기", "싱크대", "세면대", "양변기", "욕조",
        "수도관", "배수관", "보일러관",
        "도어락", "번호키", "디지털도어락",
        "냉난방기", "에어컨", "온수기",
        "폐가구", "폐기물", "고철", "건축폐기물", "사업장폐기물",
        "유리창", "강화유리", "거울",
        "샷시", "샤시", "강화도어", "자바라",
        "녹화기", "카메라",
        "노트북", "데스크탑", "윈도우",
    ],
    "place": [
        "사무실", "사무", "가정", "가정집", "아파트", "빌라", "주택", "오피스텔",
        "상가", "상점", "매장", "점포", "공장", "창고",
        "화장실", "주방", "베란다", "발코니", "옥상", "지하",
        "현관", "방", "거실",
        "병원", "학교", "교회", "성당", "절",
        "건물", "빌딩", "시설",
    ],
    "brand": [
        "LG", "엘지", "LX", "하우시스",
        "삼성", "KCC",
        "린나이", "경동", "귀뚜라미",
        "현대", "대우",
        "스카이라이프", "KT", "SK",
        "농협",
        "일성", "이건",
    ],
    "tag": [
        "24시", "24시간", "무료", "전문", "업체", "센터", "사무소",
        "당일", "즉시", "긴급", "응급",
        "최저가", "저렴", "할인",
        "1급", "특급",
        "전국", "지역",
        "AS", "A/S",
        "정품", "신품", "중고",
        "전문가", "마스터",
        "허가", "면허", "자격",
        "친절", "안전", "정확",
    ],
}


# ──────────────────────────────────────────────────────────────────────────
# Excel 로더
# ──────────────────────────────────────────────────────────────────────────
def _load_business_names() -> List[Tuple[str, float]]:
    if not BUSINESS_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(BUSINESS_XLSX, read_only=True, data_only=True)
    ws = wb.active
    out: List[Tuple[str, float]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        try:
            weight = float(row[1] or 1.0)
        except (TypeError, ValueError):
            weight = 1.0
        if name:
            out.append((name, max(weight, 1.0)))
    wb.close()
    return out


def _load_categories() -> List[Tuple[str, float]]:
    if not CATEGORY_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(CATEGORY_XLSX, read_only=True, data_only=True)
    ws = wb.active
    out: List[Tuple[str, float]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cat = str(row[0]).strip()
        try:
            weight = float(row[1] or 1.0)
        except (TypeError, ValueError):
            weight = 1.0
        if cat:
            out.append((cat, max(weight, 1.0)))
    wb.close()
    return out


# ──────────────────────────────────────────────────────────────────────────
# Auto extraction — n-gram 빈도 기반 + 최장일치 우선 가지치기
# ──────────────────────────────────────────────────────────────────────────
_BLOCK_RE = re.compile(r"[가-힣A-Za-z0-9]+")


def _extract_ngrams(names: List[Tuple[str, float]],
                    min_n: int = 2,
                    max_n: int = 6,
                    min_df: int = 3) -> Tuple[Counter, Counter]:
    """Return (document_frequency, weighted_frequency) Counters."""
    df = Counter()
    wf = Counter()
    for name, w in names:
        seen: set[str] = set()
        for block in _BLOCK_RE.findall(name):
            for n in range(min_n, max_n + 1):
                for i in range(len(block) - n + 1):
                    seen.add(block[i : i + n])
        for tok in seen:
            df[tok] += 1
            wf[tok] += w
    df = Counter({k: v for k, v in df.items() if v >= min_df})
    wf = Counter({k: v for k, v in wf.items() if k in df})
    return df, wf


def _prune_substrings(df: Counter, ratio: float = 1.2) -> List[str]:
    """Drop a token if a +1-char extension covers ≥ df(tok)/ratio occurrences."""
    by_len: Dict[int, List[str]] = defaultdict(list)
    for t in df:
        by_len[len(t)].append(t)
    drop: set[str] = set()
    for L in sorted(by_len.keys()):
        if L + 1 not in by_len:
            continue
        for ltok in by_len[L + 1]:
            ldf = df[ltok]
            for i in range(len(ltok) - L + 1):
                stok = ltok[i : i + L]
                if stok in df and df[stok] <= ldf * ratio:
                    drop.add(stok)
    return [t for t in df if t not in drop]


# ──────────────────────────────────────────────────────────────────────────
# Auto-labeling — 카테고리 추론 (시드 부분일치 + 어미 휴리스틱)
# ──────────────────────────────────────────────────────────────────────────
ACTION_SUFFIX = ("막힘", "뚫음", "설치", "수리", "교체", "공사", "시공", "철거",
                 "청소", "정리", "수거", "처리", "매입", "판매", "출장",
                 "복구", "탐지", "역류", "해빙", "동파", "이전", "조사", "찾기")
PLACE_SUFFIX = ("사무실", "가정", "아파트", "빌라", "상가", "공장", "주택",
                "화장실", "주방", "옥상", "베란다", "지하")
TAG_PREFIX = ("24시", "1급", "특급", "당일", "긴급", "응급", "무료", "전문")


def _seed_lookup() -> Dict[str, str]:
    """token -> category"""
    out: Dict[str, str] = {}
    # 시드는 우선순위가 높음 — 충돌 시 main > action > material > place > brand > tag
    for cat in CATEGORIES:
        for tok in SEED.get(cat, []):
            out.setdefault(tok, cat)
    return out


def _infer_category(tok: str, seed: Dict[str, str]) -> str:
    """규칙 기반 카테고리 추론 (AI 미사용)."""
    if tok in seed:
        return seed[tok]

    # 시드 정확 일치 우선 — 이미 처리됨
    # 1) 동작 어미 (가장 대표적인 식별 신호)
    for s in ACTION_SUFFIX:
        if tok.endswith(s):
            # 단, '하수구막힘' 같은 복합어는 main+action 양쪽에 해당하나 action 우선 (사용자 의도)
            return "action"
    # 2) 장소 어미
    for s in PLACE_SUFFIX:
        if tok.endswith(s):
            return "place"
    # 3) 태그 prefix
    for p in TAG_PREFIX:
        if tok.startswith(p):
            return "tag"
    # 4) 시드 부분일치 → main 후보
    for cat in ("main", "material"):
        for seed_tok in SEED.get(cat, []):
            if seed_tok in tok or tok in seed_tok:
                return cat
    # 5) 영문 대문자 + 한글 → brand 추정
    if re.search(r"[A-Z]{2,}", tok):
        return "brand"
    # 6) 기본 → main (명사형 후보로 간주)
    return "main"


# ──────────────────────────────────────────────────────────────────────────
# 메인 빌더
# ──────────────────────────────────────────────────────────────────────────
def build_dictionary(write: bool = True) -> Dict:
    names = _load_business_names()
    cats = _load_categories()

    df, wf = _extract_ngrams(names, min_n=2, max_n=6, min_df=3)
    kept = _prune_substrings(df, ratio=1.2)

    # 카테고리에서 추출한 atoms 추가
    cat_atoms: Counter = Counter()
    for raw, w in cats:
        for atom in re.split(r"[-/,()&\s]+", raw):
            if atom and len(atom) >= 2:
                cat_atoms[atom] += w

    seed_map = _seed_lookup()

    tokens: Dict[str, Dict] = {}

    # 1) 시드는 모두 등록 (df=0이어도 OK)
    for cat, lst in SEED.items():
        for tok in lst:
            if tok in tokens:
                continue
            tokens[tok] = {
                "category": cat,
                "df": int(df.get(tok, 0)),
                "weight": float(wf.get(tok, 0.0)),
                "source": "seed",
            }

    # 2) 카테고리 atom 추가 (main 으로 분류)
    for atom, w in cat_atoms.items():
        if atom in tokens:
            continue
        tokens[atom] = {
            "category": _infer_category(atom, seed_map),
            "df": int(df.get(atom, 0)),
            "weight": float(w),
            "source": "category",
        }

    # 3) 자동 추출 토큰 추가
    for tok in kept:
        if tok in tokens:
            # df/weight 만 갱신
            tokens[tok]["df"] = int(df[tok])
            tokens[tok]["weight"] = float(wf[tok])
            continue
        tokens[tok] = {
            "category": _infer_category(tok, seed_map),
            "df": int(df[tok]),
            "weight": float(wf[tok]),
            "source": "auto",
        }

    out = {
        "categories": CATEGORIES,
        "tokens": tokens,
        "stats": {
            "business_count": len(names),
            "category_count": len(cats),
            "total_weight": float(sum(w for _, w in names)),
            "token_count": len(tokens),
            "by_category": {
                c: sum(1 for v in tokens.values() if v["category"] == c)
                for c in CATEGORIES
            },
        },
    }

    if write:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with DICTIONARY_JSON.open("w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False)

    return out


@lru_cache(maxsize=1)
def load_dictionary() -> Dict:
    """캐시 로드 — 없으면 자동 빌드."""
    if DICTIONARY_JSON.exists():
        try:
            with DICTIONARY_JSON.open("r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return build_dictionary(write=True)


@lru_cache(maxsize=1)
def load_business_names() -> List[Tuple[str, float]]:
    return _load_business_names()


@lru_cache(maxsize=1)
def load_categories_list() -> List[Tuple[str, float]]:
    return _load_categories()
